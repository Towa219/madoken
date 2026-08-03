# 魔法の全一覧を Excel ファイルにする。
#
#   npx tsx tools/spell_list.ts   … 先に元データ(spell_list.json)を作る
#   python tools/spell_list_xlsx.py
#
# 出力: tools/魔法一覧.xlsx
#
# 見出しを固定し、絞り込み(オートフィルタ)を付けてあるので、
# Excel側で「系統=爆裂系だけ」「魔導値の高い順」といった見方ができる。

import io
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, 'spell_list.json')
OUT = os.path.join(HERE, '魔法一覧.xlsx')

# 列ごとの幅。長い説明だけ広く取る
WIDTHS = {
    '構成': 14, '素材数': 7, '種類数': 7,
    '魔法名': 30, '系統': 22, '種類': 8, '属性': 6,
    '魔導値': 8, '威力': 7, '詠唱秒': 8, '消費MP': 8, '再使用秒': 9,
    '弾速': 7, '会心率': 7, '全体対象': 9, '自傷': 6,
    '効果': 70,
}
ELEMS = ['火', '水', '風', '土', '雷', '氷', '光', '闇']

# エレメントごとの色(ゲーム内の色に寄せる)
ELEM_FILL = {
    '火': 'FFE0D6', '水': 'D6ECFF', '風': 'D9F5E4', '土': 'F0E2CC',
    '雷': 'FFF6CC', '氷': 'DDF6FF', '光': 'FFFBDD', '闇': 'EADFF7',
}


def main():
    if not os.path.exists(SRC):
        print('元データがありません。先に npx tsx tools/spell_list.ts を実行してください。')
        sys.exit(1)
    with io.open(SRC, encoding='utf-8') as f:
        doc = json.load(f)
    rows = doc['rows']
    if not rows:
        print('中身が空です。')
        sys.exit(1)

    cols = list(rows[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = '魔法一覧'

    # 1行目に作成条件を書いておく。あとで見た時に何の値か分かるように。
    ws.cell(row=1, column=1,
            value=f"魔導研究記 魔法一覧 / {doc['作成日']} / {doc['条件']} / "
                  f"{doc['件数']}件 / 系統{doc['系統数']}種")
    ws.cell(row=1, column=1).font = Font(bold=True, size=11)

    head_fill = PatternFill('solid', fgColor='4A3A78')
    head_font = Font(bold=True, color='FFFFFF')
    for i, name in enumerate(cols, start=1):
        c = ws.cell(row=2, column=i, value=name)
        c.fill = ELEM_FILL.get(name) and PatternFill('solid', fgColor=ELEM_FILL[name]) or head_fill
        c.font = Font(bold=True) if name in ELEM_FILL else head_font
        c.alignment = Alignment(horizontal='center')

    for r, row in enumerate(rows, start=3):
        for i, name in enumerate(cols, start=1):
            v = row.get(name, '')
            cell = ws.cell(row=r, column=i, value=v)
            if name in ELEMS and v == 0:
                cell.value = None          # 0は空欄にして見やすくする
            if name in ('効果', '魔法名', '系統', '構成'):
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='center')

    for i, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(name, 6)

    last = get_column_letter(len(cols))
    ws.auto_filter.ref = f'A2:{last}{len(rows) + 2}'
    ws.freeze_panes = 'A3'   # 見出しを固定して、下へ辿っても列名が見える

    wb.save(OUT)
    print(f'{OUT}  {len(rows)}件 × {len(cols)}列')


if __name__ == '__main__':
    main()
